import logging
from dataclasses import dataclass
from pathlib import Path
from typing import List

import click
from linkml_runtime.utils.schemaview import SchemaView
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from linkml._version import __version__
from linkml.utils.generator import Generator, shared_arguments


@dataclass
class ExcelGenerator(Generator):
    # ClassVars
    generatorname = Path(__file__).name
    generatorversion = "0.1.1"
    valid_formats = ["xlsx"]
    uses_schemaloader = False
    requires_metamodel = False

    split_workbook_by_class: bool = False
    include_mixins: bool = False

    def __post_init__(self) -> None:
        super().__post_init__()
        self.logger = logging.getLogger(__name__)
        self.schemaview = SchemaView(self.schema)

    @staticmethod
    def create_workbook(workbook_path: Path) -> Workbook:
        """
        Creates an Excel workbook using the openpyxl library and returns it.

        :param workbook_path: Path of the workbook to be created.
        :return: An openpyxl Workbook object representing the newly created workbook.
        """
        workbook = Workbook()
        workbook.save(workbook_path)
        return workbook

    def create_workbook_and_worksheets(self, output_path: Path, classes: List[str]) -> None:
        """
        Creates a workbook with worksheets for each class.

        :param output_path: The path where the workbook should be created.
        :param classes: List of class names for which worksheets should be created.
        """
        workbook = self.create_workbook(output_path)
        workbook.remove(workbook.active)
        sv = self.schemaview

        for cls_name in classes:
            workbook.create_sheet(cls_name)

            # Add columns to the worksheet for the current class
            slots = [s.name for s in sv.class_induced_slots(cls_name, self.mergeimports)]
            self.add_columns_to_worksheet(workbook, cls_name, slots)
            workbook.save(output_path)

            # Add enum validation for columns with enum types
            enum_list = list(sv.all_enums(imports=self.mergeimports).keys())
            for s in sv.class_induced_slots(cls_name, self.mergeimports):
                if s.range in enum_list:
                    pv_list = list(sv.get_enum(s.range).permissible_values.keys())

                    # Data Validation formula to be applied to the column and
                    # which will be used to create the dropdown
                    dv_formula = f'"{",".join(pv_list)}"'

                    # Check if the total length of the data validation formula
                    # including the separators is <= 255 characters
                    enum_length = len(dv_formula)
                    if enum_length <= 255:
                        self.column_enum_validation(workbook, cls_name, s.name, dv_formula)
                    else:
                        self.logger.warning(
                            f"'{s.range}' has permissible values with total "
                            "length > 255 characters. Dropdowns may not work properly "
                            f"in {output_path}"
                        )
                workbook.save(output_path)

        workbook.save(output_path)
        if self.split_workbook_by_class:
            self.logger.info(f"The Excel workbooks have been written to {output_path}")

    @staticmethod
    def add_columns_to_worksheet(workbook: Workbook, worksheet_name: str, sheet_headings: List[str]) -> None:
        """
        Get a worksheet by name and add a column to it in an existing workbook.

        :param workbook: The workbook to which the worksheet should be added.
        :param worksheet_name: Name of the worksheet to add the column to.
        :param column_data: List of data to populate the column with.
        """
        # Get the worksheet by name
        worksheet = workbook[worksheet_name]

        # Add the headings to the worksheet
        for i, heading in enumerate(sheet_headings):
            worksheet.cell(row=1, column=i + 1, value=heading)

    @staticmethod
    def column_enum_validation(
        workbook: Workbook,
        worksheet_name: str,
        column_name: str,
        dv_formula: str,
    ) -> None:
        """
        Get worksheet by name and add a dropdown to a specific column in it
        based on a list of values.

        :param workbook: The workbook to which the worksheet should be added.
        :param worksheet_name: Name of the worksheet to add the column dropdown to.
        :param column_name: Name of the worksheet column to add the dropdown to.
        :param dv_formula: Validation formula (as a literal) to be used for the dropdown.
        """
        worksheet = workbook[worksheet_name]

        column_list = [cell.value for cell in worksheet[1]]
        column_number = column_list.index(column_name) + 1
        column_letter = get_column_letter(column_number)

        # Create the data validation object and set the dropdown values
        dv = DataValidation(type="list", formula1=dv_formula, allow_blank=True)

        worksheet.add_data_validation(dv)

        dv.add(f"{column_letter}2:{column_letter}1048576")

    def serialize(self, **kwargs) -> str:
        sv = self.schemaview

        if self.include_mixins:
            classes_to_process = [
                cls_name for cls_name, cls in sv.all_classes(imports=self.mergeimports).items() if not cls.abstract
            ]
        else:
            classes_to_process = [
                cls_name
                for cls_name, cls in sv.all_classes(imports=self.mergeimports).items()
                if not cls.mixin and not cls.abstract
            ]

        if self.split_workbook_by_class:
            output_path = Path(self.schema.name + "_worksheets") if not self.output else Path(self.output)
            output_path = output_path.absolute()

            if not output_path.is_dir():
                output_path.mkdir(parents=True, exist_ok=True)

            for cls_name in classes_to_process:
                cls_output_path = output_path.joinpath(f"{cls_name}.xlsx")
                self.create_workbook_and_worksheets(cls_output_path, [cls_name])
                self.logger.info(f"The Excel workbook for class '{cls_name}' has been written to {cls_output_path}")
        else:
            output_path = Path(self.schema.name + ".xlsx") if not self.output else Path(self.output)
            output_path = output_path.absolute()

            self.create_workbook_and_worksheets(output_path, classes_to_process)
            self.logger.info(f"The Excel workbook has been written to {output_path}")


@shared_arguments(ExcelGenerator)
@click.command()
@click.option(
    "--split-workbook-by-class",
    is_flag=True,
    default=False,
    help="""Split model into separate Excel workbooks/files, one for each class""",
)
@click.option(
    "--include-mixins",
    is_flag=True,
    default=False,
    help="""Include mixin classes in the generated Excel workbook/workbooks""",
)
@click.option(
    "-o",
    "--output",
    type=click.Path(),
    help="""Name of Excel spreadsheet to be created, or name of directory to create split workbooks in""",
)
@click.version_option(__version__, "-V", "--version")
def cli(yamlfile, split_workbook_by_class, include_mixins, **kwargs):
    """Generate Excel representation of a LinkML model"""
    ExcelGenerator(
        yamlfile, split_workbook_by_class=split_workbook_by_class, include_mixins=include_mixins, **kwargs
    ).serialize(**kwargs)


if __name__ == "__main__":
    cli()
