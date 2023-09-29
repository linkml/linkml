from openpyxl import load_workbook

from linkml.generators.excelgen import ExcelGenerator

# path to example organization schema


def test_excel_generation(input_path, tmp_path):
    organization_schema = str(input_path("organization.yaml"))
    xlsx_filename = str(tmp_path / "schema.xlsx")

    # call ExcelGenerator generator class
    ExcelGenerator(organization_schema, output=xlsx_filename).serialize()

    # load Excel workbook object from temporary file path
    wb_obj = load_workbook(xlsx_filename)

    # check the names of the created worksheets that are part of the workbook
    assert wb_obj.sheetnames == ["organization", "employee", "manager"]

    # test case to check the column names in Employee worksheet
    employee_cols_list = []
    max_col = wb_obj["employee"].max_column

    for i in range(1, max_col + 1):
        cell_obj = wb_obj["employee"].cell(row=1, column=i)
        employee_cols_list.append(cell_obj.value)

    assert sorted(employee_cols_list) == [
        "age in years",
        "aliases",
        "first name",
        "id",
        "last name",
    ]

    # test case to check the column names in Manager worksheet
    manager_cols_list = []
    max_col = wb_obj["manager"].max_column

    for i in range(1, max_col + 1):
        cell_obj = wb_obj["manager"].cell(row=1, column=i)
        manager_cols_list.append(cell_obj.value)

    assert sorted(manager_cols_list) == [
        "age in years",
        "aliases",
        "first name",
        "has employees",
        "id",
        "last name",
    ]

    # test case to check the column names in Organization worksheet
    organization_cols_list = []
    max_col = wb_obj["organization"].max_column

    for i in range(1, max_col + 1):
        cell_obj = wb_obj["organization"].cell(row=1, column=i)
        organization_cols_list.append(cell_obj.value)

    assert sorted(organization_cols_list) == ["has boss", "id", "name"]
