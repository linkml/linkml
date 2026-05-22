import hashlib
import logging
import os

import pytest
from click.testing import CliRunner
from openpyxl import load_workbook

from linkml.generators.excelgen import ExcelGenerator, cli


def test_excel_generation(input_path, tmp_path):
    organization_schema = str(input_path("organization.yaml"))
    xlsx_filename = str(tmp_path / "schema.xlsx")

    # call ExcelGenerator generator class
    ExcelGenerator(organization_schema, output=xlsx_filename).serialize()

    # load Excel workbook object from temporary file path
    wb_obj = load_workbook(xlsx_filename)

    # check the names of the created worksheets that are part of the workbook
    assert wb_obj.sheetnames == ["organization", "employee", "manager"]

    # test case to check the column names in Employee worksheet
    employee_cols_list = []
    max_col = wb_obj["employee"].max_column

    for i in range(1, max_col + 1):
        cell_obj = wb_obj["employee"].cell(row=1, column=i)
        employee_cols_list.append(cell_obj.value)

    assert sorted(employee_cols_list) == [
        "age in years",
        "aliases",
        "first name",
        "id",
        "last name",
    ]

    # test case to check the column names in Manager worksheet
    manager_cols_list = []
    max_col = wb_obj["manager"].max_column

    for i in range(1, max_col + 1):
        cell_obj = wb_obj["manager"].cell(row=1, column=i)
        manager_cols_list.append(cell_obj.value)

    assert sorted(manager_cols_list) == [
        "age in years",
        "aliases",
        "first name",
        "has employees",
        "id",
        "last name",
    ]

    # test case to check the column names in Organization worksheet
    organization_cols_list = []
    max_col = wb_obj["organization"].max_column

    for i in range(1, max_col + 1):
        cell_obj = wb_obj["organization"].cell(row=1, column=i)
        organization_cols_list.append(cell_obj.value)

    assert sorted(organization_cols_list) == ["has boss", "id", "name"]


def test_multiple_excel_workbooks_generation(input_path, tmp_path):
    organization_schema = str(input_path("organization.yaml"))
    excel_files_folder = str(tmp_path)

    ExcelGenerator(organization_schema, split_workbook_by_class=True, output=excel_files_folder).serialize()

    _, _, files = next(os.walk(excel_files_folder))
    file_count = len(files)

    assert file_count == 3

    expected_files = ["employee.xlsx", "manager.xlsx", "organization.xlsx"]
    assert sorted(files) == expected_files

    def verify_workbook_columns(excel_file_path, worksheet_name):
        wb_obj = load_workbook(excel_file_path)
        worksheet = wb_obj[worksheet_name]

        column_names = []
        max_col = worksheet.max_column

        for i in range(1, max_col + 1):
            cell_obj = worksheet.cell(row=1, column=i)
            column_names.append(cell_obj.value)

        return column_names

    for file in files:
        file_path = os.path.join(excel_files_folder, file)
        wb_obj = load_workbook(file_path)
        worksheet_count = len(wb_obj.sheetnames)
        assert worksheet_count == 1

        if file == "employee.xlsx":
            column_names = verify_workbook_columns(file_path, "employee")
            assert sorted(column_names) == [
                "age in years",
                "aliases",
                "first name",
                "id",
                "last name",
            ]

        if file == "manager.xlsx":
            column_names = verify_workbook_columns(file_path, "manager")
            assert sorted(column_names) == [
                "age in years",
                "aliases",
                "first name",
                "has employees",
                "id",
                "last name",
            ]

        if file == "organization.xlsx":
            column_names = verify_workbook_columns(file_path, "organization")
            assert sorted(column_names) == ["has boss", "id", "name"]


def test_file_generation_for_mixins(input_path, tmp_path):
    kitchen_sink_schema = str(input_path("kitchen_sink.yaml"))
    excel_files_folder = str(tmp_path)

    # Call ExcelGenerator generator class without including mixins
    ExcelGenerator(kitchen_sink_schema, split_workbook_by_class=True, output=excel_files_folder).serialize()

    _, _, files = next(os.walk(excel_files_folder))
    file_count = len(files)

    # Check that HasAliases.xlsx and WithLocation.xlsx have not been generated
    assert file_count == 31

    assert "HasAliases.xlsx" not in files
    assert "WithLocation.xlsx" not in files

    # Call ExcelGenerator generator class with mixins
    ExcelGenerator(
        kitchen_sink_schema, split_workbook_by_class=True, output=excel_files_folder, include_mixins=True
    ).serialize()

    _, _, files = next(os.walk(excel_files_folder))
    file_count = len(files)

    # Check if HasAliases.xlsx and WithLocation.xlsx have been generated
    assert file_count == 33

    assert "HasAliases.xlsx" in files
    assert "WithLocation.xlsx" in files


def test_serialize_empty_schema_raises_error(tmp_path):
    """Test that serialize raises ValueError when schema has no classes defined."""
    schema_yaml = """
id: https://example.org/empty
name: empty_schema
prefixes:
  linkml: https://w3id.org/linkml/

classes:
"""
    schema_file = tmp_path / "empty_schema.yaml"
    schema_file.write_text(schema_yaml)
    xlsx_filename = tmp_path / "empty_schema.xlsx"

    generator = ExcelGenerator(str(schema_file), output=str(xlsx_filename))

    with pytest.raises(ValueError, match="No classes defined in the schema"):
        generator.serialize()

    assert not xlsx_filename.exists()


def test_only_abstract_classes_raises_error(tmp_path):
    """Test that Excel generator raises ValueError when schema has only abstract classes."""
    schema_yaml = """
id: https://example.org/abstract-only
name: abstract_only
prefixes:
  linkml: https://w3id.org/linkml/

classes:
  AbstractBase:
    abstract: true
    attributes:
      id:
        range: string
  AnotherAbstract:
    abstract: true
    attributes:
      name:
        range: string
"""
    schema_file = tmp_path / "abstract_only.yaml"
    schema_file.write_text(schema_yaml)
    xlsx_filename = tmp_path / "abstract_only.xlsx"

    generator = ExcelGenerator(str(schema_file), output=str(xlsx_filename))

    with pytest.raises(ValueError, match="Schema contains only abstract classes.*AbstractBase.*AnotherAbstract"):
        generator.serialize()

    assert not xlsx_filename.exists()


def test_only_mixin_classes_raises_error(tmp_path):
    """Test that Excel generator raises ValueError when schema has only mixin classes."""
    schema_yaml = """
id: https://example.org/mixin-only
name: mixin_only
prefixes:
  linkml: https://w3id.org/linkml/

classes:
  MixinA:
    mixin: true
    attributes:
      field_a:
        range: string
  MixinB:
    mixin: true
    attributes:
      field_b:
        range: string
"""
    schema_file = tmp_path / "mixin_only.yaml"
    schema_file.write_text(schema_yaml)
    xlsx_filename = tmp_path / "mixin_only.xlsx"

    generator = ExcelGenerator(str(schema_file), output=str(xlsx_filename))

    with pytest.raises(ValueError, match="Schema contains only mixin classes.*--include-mixins flag"):
        generator.serialize()

    assert not xlsx_filename.exists()


def test_only_mixin_classes_with_include_mixins(tmp_path):
    """Test that mixin-only schema works when --include-mixins is used."""
    schema_yaml = """
id: https://example.org/mixin-only
name: mixin_only
prefixes:
  linkml: https://w3id.org/linkml/

classes:
  MixinA:
    mixin: true
    attributes:
      field_a:
        range: string
"""
    schema_file = tmp_path / "mixin_only.yaml"
    schema_file.write_text(schema_yaml)
    xlsx_filename = tmp_path / "mixin_only.xlsx"

    generator = ExcelGenerator(str(schema_file), output=str(xlsx_filename), include_mixins=True)
    generator.serialize()

    # File should be created when include_mixins is True
    assert xlsx_filename.exists()


def test_abstract_and_mixin_classes_only_raises_error(tmp_path):
    """Test that Excel generator raises ValueError when schema has only abstract and mixin classes."""
    schema_yaml = """
id: https://example.org/abstract-mixin
name: abstract_mixin
prefixes:
  linkml: https://w3id.org/linkml/

classes:
  AbstractBase:
    abstract: true
    attributes:
      id:
        range: string
  MixinHelper:
    mixin: true
    attributes:
      helper_field:
        range: string
"""
    schema_file = tmp_path / "abstract_mixin.yaml"
    schema_file.write_text(schema_yaml)
    xlsx_filename = tmp_path / "abstract_mixin.xlsx"

    generator = ExcelGenerator(str(schema_file), output=str(xlsx_filename))

    with pytest.raises(ValueError, match="abstract classes.*mixin classes.*--include-mixins flag"):
        generator.serialize()

    assert not xlsx_filename.exists()


def test_abstract_and_mixin_with_include_mixins(tmp_path):
    """Test that abstract+mixin schema creates file for mixins when --include-mixins is used."""
    schema_yaml = """
id: https://example.org/abstract-mixin
name: abstract_mixin
prefixes:
  linkml: https://w3id.org/linkml/

classes:
  AbstractBase:
    abstract: true
    attributes:
      id:
        range: string
  MixinHelper:
    mixin: true
    attributes:
      helper_field:
        range: string
"""
    schema_file = tmp_path / "abstract_mixin.yaml"
    schema_file.write_text(schema_yaml)
    xlsx_filename = tmp_path / "abstract_mixin.xlsx"

    generator = ExcelGenerator(str(schema_file), output=str(xlsx_filename), include_mixins=True)
    generator.serialize()

    # File should be created for the mixin (abstract still excluded)
    assert xlsx_filename.exists()


def test_enum_validation_adds_dropdown(input_path, tmp_path):
    """Test that slots that have an enum range asserted on them get a
    DataValidation dropdown in the worksheet.
    """
    kitchen_sink_schema = str(input_path("kitchen_sink.yaml"))
    xlsx_filename = tmp_path / "kitchen_sink_enum.xlsx"
    ExcelGenerator(kitchen_sink_schema, output=str(xlsx_filename)).serialize()

    wb = load_workbook(xlsx_filename)
    assert "Person" in wb.sheetnames

    ws = wb["FamilialRelationship"]
    # Column headers must include both slots
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    assert "cordialness" in headers
    assert "type" in headers

    # The DataValidation list for the enum column must be present
    dv_formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("SIBLING_OF" in f and "PARENT_OF" in f and "CHILD_OF" in f for f in dv_formulas)


def test_enum_values_exceed_255_chars_logs_warning(input_path, tmp_path, caplog):
    """Test that a warning is logged when enum permissible values exceed 255 characters."""

    kitchen_sink_schema = str(input_path("kitchen_sink.yaml"))
    xlsx_filename = tmp_path / "kitchen_sink_enum.xlsx"
    ExcelGenerator(kitchen_sink_schema, output=str(xlsx_filename)).serialize()

    # https://docs.pytest.org/en/stable/how-to/logging.html#caplog-fixture
    with caplog.at_level(logging.WARNING, logger="linkml.generators.excelgen"):
        ExcelGenerator(kitchen_sink_schema, output=str(xlsx_filename)).serialize()

    # Assert a warning was logged about enum values exceeding 255 characters
    assert any("255" in record.message and "LongEnum" in record.message for record in caplog.records)


##### tests for long class names (> 31 chars)

LONG_NAME_SCHEMA = """\
id: https://example.org/excelgen-longname
name: excelgen_longname
prefixes:
  linkml: https://w3id.org/linkml/
  ex: https://example.org/
default_prefix: ex
imports:
  - linkml:types

classes:
  AClassNameThatIsDefinitelyLongerThan31Chars:
    description: 43-char class name to overflow Excel 31-char sheet title limit.
    attributes:
      id:
        identifier: true
        range: string
      label:
        range: string
  ShortName:
    attributes:
      id:
        identifier: true
        range: string
"""


def test_long_class_name_does_not_crash(tmp_path):
    """gen-excel must not raise KeyError when a class name exceeds 31 characters."""
    schema_path = tmp_path / "long_name.yaml"
    schema_path.write_text(LONG_NAME_SCHEMA)
    xlsx_path = tmp_path / "out.xlsx"

    ExcelGenerator(str(schema_path), output=str(xlsx_path)).serialize()

    wb = load_workbook(xlsx_path)

    # All sheet titles must be at most 31 characters
    assert all(len(name) <= 31 for name in wb.sheetnames)

    # ShortName fits unchanged
    assert "ShortName" in wb.sheetnames

    # The long class sheet title must be a deterministic 31-char truncation
    long_cls = "AClassNameThatIsDefinitelyLongerThan31Chars"
    h = hashlib.sha1(long_cls.encode("utf-8")).hexdigest()[:6]
    expected_title = f"{long_cls[:24]}_{h}"
    assert expected_title in wb.sheetnames

    # The truncated sheet must carry the expected column headers
    ws = wb[expected_title]
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    assert "id" in headers
    assert "label" in headers


def test_long_class_name_logs_warning(tmp_path, caplog):
    """A warning must be emitted when a class name is truncated."""
    schema_path = tmp_path / "long_name.yaml"
    schema_path.write_text(LONG_NAME_SCHEMA)
    xlsx_path = tmp_path / "out.xlsx"

    with caplog.at_level(logging.WARNING, logger="linkml.generators.excelgen"):
        ExcelGenerator(str(schema_path), output=str(xlsx_path)).serialize()

    long_cls = "AClassNameThatIsDefinitelyLongerThan31Chars"
    assert any(long_cls in record.message for record in caplog.records)


def test_safe_sheet_title_short_name():
    """Names ≤ 31 chars pass through unchanged."""
    used: set = set()
    assert ExcelGenerator._safe_sheet_title("ShortName", used) == "ShortName"
    assert "ShortName" in used


def test_safe_sheet_title_long_name():
    """Names > 31 chars are truncated to exactly 31 chars."""
    used: set = set()
    long_cls = "AClassNameThatIsDefinitelyLongerThan31Chars"
    title = ExcelGenerator._safe_sheet_title(long_cls, used)
    assert len(title) == 31
    h = hashlib.sha1(long_cls.encode("utf-8")).hexdigest()[:6]
    assert title == f"{long_cls[:24]}_{h}"


def test_safe_sheet_title_collision_disambiguation():
    """Collisions after truncation are resolved deterministically."""
    used: set = set()

    # Two distinct long names sharing same 24-char prefix have different titles.
    cls_a = "AClassNameThatIsDefinitelyLongerThan31Chars"
    cls_b = "AClassNameThatIsDefinitelyLongerThan31CharsB"
    title_a = ExcelGenerator._safe_sheet_title(cls_a, used)
    title_b = ExcelGenerator._safe_sheet_title(cls_b, used)
    assert title_a != title_b
    assert len(title_a) <= 31
    assert len(title_b) <= 31


def test_safe_sheet_title_collision_loop_executes():
    """The while-loop inside _safe_sheet_title fires when the candidate title
    is already occupied."""
    used: set = {"ShortName"}  # pre-populate so the first candidate collides
    title = ExcelGenerator._safe_sheet_title("ShortName", used)
    # The loop must have run: result differs from the original and stays ≤ 31
    assert title != "ShortName"
    assert len(title) <= 31
    # The disambiguated title must now be in used
    assert title in used
    assert "ShortName" in used  # original still present


def test_split_workbook_creates_output_dir(input_path, tmp_path):
    """serialize() with split_workbook_by_class=True must create the output
    directory when it does not yet exist."""
    organization_schema = str(input_path("organization.yaml"))
    new_dir = tmp_path / "new_subdir" / "sheets"  # does not exist

    ExcelGenerator(
        organization_schema,
        split_workbook_by_class=True,
        output=str(new_dir),
    ).serialize()

    assert new_dir.is_dir()
    xlsx_files = list(new_dir.glob("*.xlsx"))
    assert len(xlsx_files) > 0


def test_cli_entry_point(input_path, tmp_path):
    """The cli() Click command must produce a workbook without errors."""
    organization_schema = str(input_path("organization.yaml"))
    xlsx_path = str(tmp_path / "cli_out.xlsx")

    runner = CliRunner()
    result = runner.invoke(cli, [organization_schema, "-o", xlsx_path])

    assert result.exit_code == 0, result.output
    wb = load_workbook(xlsx_path)
    assert len(wb.sheetnames) > 0
